from io import BytesIO
from sqlalchemy import create_engine
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from zipfile import ZipFile

class DatabaseHandler:

    def __init__(self):
        self.engine = None

    def connect(self, DB_SETTING):
        try:
            db_url = f"postgresql://{DB_SETTING['user']}:{DB_SETTING['password']}@" \
                     f"{DB_SETTING['host']}:{DB_SETTING['port']}/{DB_SETTING['database']}"
            self.engine = create_engine(db_url)
            print("Подключение к базе данных успешно выполнено.")
        except Exception as e:
            print(f"Ошибка подключения к базе данных: {e}")

    def fetch_data(self, query):
        if self.engine:
            try:
                with self.engine.connect() as connection:
                    return pd.read_sql_query(query, connection)
            except Exception as e:
                print(f"Ошибка при выполнении запроса: {e}")
                return None
        else:
            print("Отсутствует активное подключение к базе данных.")
            return None

    def close_connection(self):
        if self.engine:
            self.engine.dispose()
            print("Подключение к базе данных закрыто.")
        else:
            print("Нет активного подключения для закрытия.")

    def format_excel(self, workbook):
        sheet = workbook.active

        max_row = sheet.max_row
        max_col = sheet.max_column
        table_range = f"A1:{sheet.cell(row=max_row, column=max_col).coordinate}"

        table = Table(displayName="Table1", ref=table_range)

        style = TableStyleInfo(
            name="TableStyleMedium13", 
            showRowStripes=True
        )
        table.tableStyleInfo = style

        sheet.add_table(table)

        font = Font(name="Times New Roman", size=14)
        alignment = Alignment(horizontal="left", vertical="center")
        header_alignment = Alignment(horizontal="center", vertical="center")  
        center_alignment_first_column = Alignment(horizontal="center", vertical="center")
        fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for cell in sheet[1]:
            cell.font = Font(bold=True, name="Times New Roman", size=14)
            cell.alignment = header_alignment  
            cell.fill = fill

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = font
                cell.alignment = alignment

        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = center_alignment_first_column

        for column in sheet.columns:
            column_letter = column[0].column_letter
            if column_letter == 'A':  
                sheet.column_dimensions[column_letter].width = 10
            else:
                sheet.column_dimensions[column_letter].width = 25

    def process_table(self, table_name, config):
        table_config = config["tables"].get(table_name)

        if not table_config:
            print(f"Таблица {table_name} отсутствует в конфиге.")
            return

        query = table_config["query"]
        fields = table_config.get("fields")  
        if not fields:
            print(f"Для таблицы '{table_name}' отсутствует список полей в конфиге.")
            return

        data = self.fetch_data(query)

        if data is not None and not data.empty:
            df = data
            if len(fields) != len(df.columns):
                print("Количество столбцов в данных и в 'fields' не совпадает.")
                return

            df.columns = fields

            for col in df.columns:
                if "дата" in col.lower() or "время" in col.lower():
                    df[col] = pd.to_datetime(df[col]).dt.strftime("%Y-%m-%d %H:%M:%S")

            output_stream = BytesIO()
            with pd.ExcelWriter(output_stream, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)
                workbook = writer.book
                self.format_excel(workbook)

            output_stream.seek(0)
            return output_stream

        else:
            print(f"Нет данных для таблицы '{table_name}'.")
            return None
        
    def process_archive(self, config):
        if not config.get("tables"):
            print("Конфигурация не содержит таблиц.")
            return None

        output_archive = BytesIO()

        with ZipFile(output_archive, 'w') as archive:
            for table_name in config["tables"].keys():
                print(f"Обработка таблицы: {table_name}")
                excel_stream = self.process_table(table_name, config)

                if excel_stream:
                    file_name = f"{table_name}.xlsx"
                    archive.writestr(file_name, excel_stream.getvalue())
                else:
                    print(f"Таблица '{table_name}' не добавлена в архив.")

        output_archive.seek(0)
        return output_archive
        
